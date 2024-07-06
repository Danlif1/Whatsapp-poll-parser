import json

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.discovery import build
from dotenv import load_dotenv
import os
import sqlite3
from typing import List
from datetime import datetime, timedelta
import blackboxprotobuf
from hexdump import hexdump
import subprocess
import hexdump

message_types = {"POLL": 46}

load_dotenv()
path_to_db = os.getenv('path_to_db')
print(path_to_db)

people_map = {}


def initalize_people():
    global people_map
    people_map = json.loads(os.getenv('people_map'))


def translate_data(byte_data):
    message, typedef = blackboxprotobuf.decode_message(byte_data)
    header = message['8']['2'].decode('utf8')
    questions = []
    for _message in message['8']['3']:
        questions.append(_message['1'].decode('utf8'))
    votes_for = [0] * len(questions)
    for _message in message['8']['5']:
        try:
            if isinstance(_message['1'], list):
                for _vote in _message['1']:
                    votes_for[_vote] += 1
            else:
                votes_for[_message['1']] += 1
        except:
            pass
    result = {"header": header, "questions": questions, "votes": votes_for}
    return result


def turn_off_wifi():
    # This command works for macOS
    subprocess.run(["networksetup", "-setairportpower", "airport", "off"], check=True)


def turn_on_wifi():
    # This command works for macOS
    subprocess.run(["networksetup", "-setairportpower", "airport", "on"], check=True)


def force_quit_whatsapp():
    """Force quit WhatsApp application on macOS"""
    try:
        subprocess.run(["pkill", "WhatsApp"], check=True)
        print("WhatsApp has been force-quit.")
    except subprocess.CalledProcessError:
        print("Failed to force-quit WhatsApp or WhatsApp is not running.")


def open_whatsapp():
    """Open WhatsApp application on macOS"""
    try:
        subprocess.run(["open", "-a", "WhatsApp"], check=True)
        print("WhatsApp has been opened.")
    except subprocess.CalledProcessError:
        print("Failed to open WhatsApp or WhatsApp is not installed.")


class Session:
    chats_table = "ZWACHATSESSION"
    messages_table = "ZWAMESSAGE"
    messages_info_table = "ZWAMESSAGEINFO"
    messages_media_table = "ZWAMEDIAITEM"
    group_members_id_table = "ZWAGROUPMEMBER"
    group_members_name_table = "ZWAPROFILEPUSHNAME"

    def __init__(self, path_to_db):
        """Connects to the db"""
        self.conn = sqlite3.connect(path_to_db)
        self.cursor = self.conn.cursor()

    class Chat:
        def __init__(self, id):
            self.id = id

    class Message:
        def __init__(self, id, cursor):
            self.id = id
            self.cursor = cursor

        def get_time(self):
            """From messages_table take ZMESSAGEDATE (Z_PK = id)"""
            self.cursor.execute(f"SELECT ZMESSAGEDATE FROM {Session.messages_table} WHERE Z_PK = ?", (self.id,))
            timestamp = self.cursor.fetchone()[0]
            if timestamp:
                original_time = datetime.utcfromtimestamp(timestamp)
                modified_time = original_time + timedelta(days=31 * 365 + 31 // 4,
                                                          hours=3)  # Adding 31 years and 3 hours
                return modified_time.strftime("%d.%m.%y %H:%M")
            return None

        def get_sender(self):
            """
           From messages_table take ZGROUPMEMBER (Z_PK = id)
           From group_members_id_table take ZMEMBERJID (Z_PK = sender_id)
           From group_members_name_table take ZPUSHNAME (ZJID = member_id)
           """
            self.cursor.execute(f"SELECT ZGROUPMEMBER FROM {Session.messages_table} WHERE Z_PK = ?", (self.id,))
            sender_id = self.cursor.fetchone()[0]
            if sender_id:
                self.cursor.execute(f"SELECT ZMEMBERJID FROM {Session.group_members_id_table} WHERE Z_PK = ?",
                                    (sender_id,))
                member_id = self.cursor.fetchone()[0]
                if member_id:
                    self.cursor.execute(f"SELECT ZPUSHNAME FROM {Session.group_members_name_table} WHERE ZJID = ?",
                                        (member_id,))
                    member_name = self.cursor.fetchone()[0]
                    return member_name
            return None

        def get_content(self):
            """From messages_table take ZTEXT (Z_PK = id)"""
            self.cursor.execute(f"SELECT ZTEXT FROM {Session.messages_table} WHERE Z_PK = ?", (self.id,))
            result = self.cursor.fetchone()[0]
            return translate_data(result) if result else None

    class Poll(Message):
        def __init__(self, id, cursor):
            super().__init__(id, cursor)

        def get_info(self, raw=False):
            """From messages_info_table take ZRECIPIENTINFO (ZMESSAGE = id)"""
            self.cursor.execute(f"SELECT ZRECEIPTINFO FROM {Session.messages_info_table} WHERE ZMESSAGE = ?",
                                (self.id,))
            result = self.cursor.fetchone()[0]
            # print(result.hex())
            if raw:
                return hexdump(result)
            return translate_data(result) if result else None

    def get_chat_id(self, chat_name) -> int:
        """
       From chats_table take Z_PK (ZPARTNERNAME = chat_name)
       Uses SQL query
       """
        self.cursor.execute(f"SELECT Z_PK FROM {Session.chats_table} WHERE ZPARTNERNAME = ?", (chat_name,))
        result = self.cursor.fetchone()
        return result[0] if result else None

    def get_messages_by_chat_id(self, chat_id) -> List[Message]:
        """
       From messages_table take all Z_PK (ZCHATSESSION = chat_id) sort by ZMESSAGEDATE (high to low)
       Uses SQL query
       """
        self.cursor.execute(
            f"SELECT Z_PK FROM {Session.messages_table} WHERE ZCHATSESSION = ? ORDER BY ZMESSAGEDATE ASC", (chat_id,))
        return [Session.Message(id=row[0], cursor=self.cursor) for row in self.cursor.fetchall()]

    def get_members_by_chat_id(self, chat_id) -> List[str]:
        """
       From group_members_id_table get all ZMEMBERJID and put into a list (ZCHATSESSION = chat_id)
       return the list
       """
        self.cursor.execute(f"SELECT ZMEMBERJID FROM {Session.group_members_id_table} WHERE ZCHATSESSION = ?",
                            (chat_id,))
        return [row[0] for row in self.cursor.fetchall()]

    def mention_everyone(self, chat_id, message_id):
        """
       Call get_members_by_chat_id(chat_id)
       Create a string called mention_string for each element in the list we get do the following:
       Append B Append the element
       Get the last message id in get_message_by_chat_id(chat_id)
       Edit that message in the db such that:
       ZTEXT would be @everyone and (Z_PK = last message id)
       Edit the row in messages_media_table where ZMESSAGE = last message id, Make ZMETADATA = mention_string
       """
        # Get members
        members = self.get_members_by_chat_id(chat_id)

        # Create mention string
        mention_string = ''.join(f'B{member}\n' for member in members)

        # Get the last message id
        messages = self.get_messages_by_chat_id(chat_id)
        last_message_id = messages[-1].id if messages else None

        if last_message_id:
            # Update the last message to mention everyone
            self.cursor.execute(f"UPDATE {Session.messages_table} SET ZTEXT = ? WHERE Z_PK = ?",
                                ('@everyone', last_message_id))

            # Update the metadata for the media item
            self.cursor.execute(f"UPDATE {Session.messages_media_table} SET ZMETADATA = ? WHERE ZMESSAGE = ?",
                                (mention_string, last_message_id))

            self.conn.commit()

    def get_messages_by_chat_id_and_type(self, chat_id, message_type) -> List[Message]:
        """
       From messages_table take all Z_PK (ZCHATSESSION = chat_id, ZMESSAGETYPE = message_type) sort by ZMESSAGEDATE (high to low)
       Uses SQL query
       """
        self.cursor.execute(
            f"SELECT Z_PK FROM {Session.messages_table} WHERE ZCHATSESSION = ? AND ZMESSAGETYPE = ? ORDER BY ZMESSAGEDATE ASC",
            (chat_id, message_type))
        if message_type == message_types["POLL"]:
            return [Session.Poll(id=row[0], cursor=self.cursor) for row in self.cursor.fetchall()]
        else:
            return [Session.Message(id=row[0], cursor=self.cursor) for row in self.cursor.fetchall()]


def extract_polls(chat_name):
    polls = []
    session = Session(path_to_db)
    my_chat = session.get_chat_id(chat_name)
    my_messages = session.get_messages_by_chat_id_and_type(my_chat, 46)
    for message in my_messages:
        poll = message.get_info(raw=False)
        poll['creator'] = message.get_sender()
        if poll['creator'] is None:
            poll['creator'] = 'דניאל'
        poll['time'] = message.get_time()
        polls.append(poll)
    return polls


def extract_poll_data(chat_name):
    poll_number = 1

    # Every line in the polls excel,
    # first line isn't from whatsapp and is the title of each column,
    # see commented code to see how to generate (bottom of the page).
    all_polls = [
        ['', 'יוצר הסקר', 'תאריך הסקר', 'כותרת הסקר', 'אופציה 1', 'מספר עונים על 1', 'אופציה 2', 'מספר עונים על 2',
         'אופציה 3', 'מספר עונים על 3', 'אופציה 4', 'מספר עונים על 4', 'אופציה 5', 'מספר עונים על 5', 'אופציה 6',
         'מספר עונים על 6', 'אופציה 7', 'מספר עונים על 7', 'אופציה 8', 'מספר עונים על 8', 'אופציה 9', 'מספר עונים על 9',
         'אופציה 10', 'מספר עונים על 10', 'אופציה 11', 'מספר עונים על 11', 'אופציה 12', 'מספר עונים על 12',
         'תאריך בדיקה אחרון', 'בודק אחרון']]
    # Every parsed poll starts with the poll number.
    single_poll = []
    single_poll.append(poll_number)
    polls = extract_polls(chat_name)
    for poll in polls:
        try:
            single_poll.append(people_map[poll['creator']])
        except:
            print("error", people_map)
            single_poll.append(poll['creator'])
        single_poll.append(poll['time'])
        single_poll.append(poll['header'])
        actual_index = 0
        for question in poll['questions']:
            single_poll.append(question)
            single_poll.append(poll['votes'][actual_index])
            actual_index += 1
        today = datetime.now()
        today = today.strftime("%d.%m.%y %H:%M")
        single_poll.append(today)
        single_poll.append("דניאל")
        all_polls.append(single_poll)
        poll_number += 1
        single_poll = []
        single_poll.append(poll_number)
    return all_polls


# Create the Excel sheet.
def create_sheet(poll_data):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Setting the font and border of each cell.
    # (13 in Excel = 10 in google sheets, thin border in Excel = border in google sheets)
    font = openpyxl.styles.Font(name='Arial', size=13)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    first = 1
    for row_index, data_list in enumerate(poll_data, start=1):
        # For each cell in the row (we have 30 cells in each row),
        # we want to update the font and border (and alignment).
        for i in range(0, 30):
            cell = ws.cell(row=row_index, column=i + 1)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='bottom', text_rotation=180)
        # If it's the first line we want to add it as normal.
        if first:
            length = len(data_list)
            for i in range(0, length):
                cell = ws.cell(row=row_index, column=i + 1, value=data_list[i])
                cell.border = thin_border
                cell.font = font
            first = 0
            continue
        length = len(data_list)
        # Otherwise we do something a little different.
        for i in range(0, length - 2):
            cell = ws.cell(row=row_index, column=i + 1, value=data_list[i])
            cell.border = thin_border
            cell.font = font
            if i % 2 != 0 and 5 <= i <= len(data_list):
                value = float(data_list[i])
                min_value = min(float(val) for val in data_list[5:len(data_list) - 2:2])
                max_value = max(float(val) for val in data_list[5:len(data_list) - 2:2])
                color = calculate_color(min_value, max_value, value)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        # The last two cells are for the date of the update and the person who updated.
        ws.cell(row=row_index, column=29, value=data_list[-2]).font = font
        ws.cell(row=row_index, column=30, value=data_list[-1]).font = font

    wb.save("poll_data.xlsx")


# color all votes.
def calculate_color(min_value, max_value, value):
    # 222,239,245 is the maximum color,
    # 35,158,208 is the minimum color.
    ratio = 0
    if min_value == max_value:
        ratio = 1
    else:
        ratio = (value - min_value) / (max_value - min_value)
    r = int((222 - 35) * (1 - ratio)) + 35
    g = int((239 - 158) * (1 - ratio)) + 158
    b = int((245 - 208) * (1 - ratio)) + 208
    return f'{r:02x}{g:02x}{b:02x}'


# transfer from excel to google sheet.
def hex_to_rgb(hex_color):
    """Convert hex color (e.g., 'FF0000') to RGB tuple (red, green, blue)."""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))


def transfer_sheet(from_name, to_name):
    # The name of the Excel file.
    excel_file = from_name
    # The id of the Google sheet.
    sheet_id = to_name

    # Credentials stuff.
    scope = ['https://spreadsheets.google.com/feeds',
             'https://www.googleapis.com/auth/drive',
             'https://www.googleapis.com/auth/spreadsheets']
    credentials = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)
    client = gspread.authorize(credentials)

    # Copying the Excel.
    df = pd.read_excel(excel_file)
    df.replace([np.inf, -np.inf], '', inplace=True)
    df.fillna('', inplace=True)

    # Updating the Google sheet.
    sheet = client.open_by_key(sheet_id)
    worksheet = sheet.get_worksheet(0)
    data = df.values.tolist()
    worksheet.update('A2', data)

    # Build the service for the Google Sheets API
    service = build('sheets', 'v4', credentials=credentials)

    # Read the Excel file with formatting
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    requests = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
        for col_idx, cell in enumerate(row):
            if cell.fill.start_color.rgb is not None:
                rgb_color = hex_to_rgb(cell.fill.start_color.rgb[2:])  # skip the first two characters 'FF'
                if rgb_color == (0, 0, 0):
                    rgb_color = (255, 255, 255)
                cell_format = {
                    'userEnteredFormat': {
                        'backgroundColor': {
                            'red': rgb_color[0] / 255.0,
                            'green': rgb_color[1] / 255.0,
                            'blue': rgb_color[2] / 255.0
                        }
                    }
                }

                requests.append({
                    'updateCells': {
                        'rows': {
                            'values': [{
                                'userEnteredFormat': cell_format['userEnteredFormat']
                            }]
                        },
                        'fields': 'userEnteredFormat.backgroundColor',
                        'start': {
                            'sheetId': worksheet.id,
                            'rowIndex': row_idx,
                            'columnIndex': col_idx
                        }
                    }
                })

    # Send batchUpdate request
    body = {
        'requests': requests
    }
    response = service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body=body).execute()

    print(f"Updated {len(response.get('replies', []))} cells with formatting.")


def main():
    # loading env.
    load_dotenv()
    chat_name = os.getenv("chat_name")
    sheet_id = os.getenv("sheet_id")
    # Creating Google sheet.
    initalize_people()
    poll_data = extract_poll_data(chat_name)
    print(poll_data)
    create_sheet(poll_data)
    # transfer_sheet("poll_data.xlsx", sheet_id)


if __name__ == "__main__":
    main()


# For the title
def split_custom_pattern(s):
    words = s.split()
    result = []
    pattern = [2, 2, 2, 2, 4, 2, 4, 2, 4, 2, 4, 2, 4, 2, 4, 2, 4, 2, 4, 2, 4, 2, 4, 2, 4, 2, 4, 3, 2]
    index = 0
    pattern_index = 0

    while index < len(words):
        # Determine the current group size based on the pattern
        group_size = pattern[pattern_index % len(pattern)]
        # Create a group of words based on the current group size
        group = words[index:index + group_size]
        # Join the group into a single string and add to the result list
        result.append(' '.join(group))
        # Move the index forward by the size of the current group
        index += group_size
        # Move to the next pattern
        pattern_index += 1

    return result

# # Example usage
# variable = "יוצר הסקר תאריך הסקר כותרת הסקר אופציה 1   מספר עונים על 1    אופציה 2   מספר עונים על 2    אופציה 3   מספר עונים על 3    אופציה 4   מספר עונים על 4    אופציה 5   מספר עונים על 5    אופציה 6   מספר עונים על 6    אופציה 7   מספר עונים על 7    אופציה 8   מספר עונים על 8    אופציה 9   מספר עונים על 9    אופציה 10  מספר עונים על 10   אופציה 11  מספר עונים על 11   אופציה 12  מספר עונים על 12   תאריך בדיקה אחרון  בודק אחרון"
# print(split_custom_pattern(variable))
