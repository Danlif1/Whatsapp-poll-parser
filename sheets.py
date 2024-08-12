import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.discovery import build


def create_sheet(poll_data):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Setting the font and border of each cell.
    # (13 in Excel = 10 in google sheets, thin border in Excel = border in google sheets)
    font = openpyxl.styles.Font(name='Arial', size=13)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    first = 1
    for row_index, data_list in enumerate(poll_data, start=1):
        # For each cell in the row (we have 30 cells in each row),
        # we want to update the font and border (and alignment).
        for i in range(0, 30):
            cell = ws.cell(row=row_index, column=i + 1)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='bottom', text_rotation=180)
        # If it's the first line we want to add it as normal.
        if first:
            length = len(data_list)
            for i in range(0, length):
                cell = ws.cell(row=row_index, column=i + 1, value=str(data_list[i]))
                cell.border = thin_border
                cell.font = font
            first = 0
            continue
        length = len(data_list)
        # Otherwise we do something a little different.
        for i in range(0, length - 2):
            cell = ws.cell(row=row_index, column=i + 1, value=str(data_list[i]))

            cell.border = thin_border
            cell.font = font
            if i % 2 != 0 and 5 <= i <= len(data_list):
                value = float(data_list[i])
                min_value = min(float(val) for val in data_list[5:len(data_list) - 2:2])
                max_value = max(float(val) for val in data_list[5:len(data_list) - 2:2])
                color = calculate_color(min_value, max_value, value)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        # The last two cells are for the date of the update and the person who updated.
        ws.cell(row=row_index, column=29, value=str(data_list[-2])).font = font
        ws.cell(row=row_index, column=30, value=str(data_list[-1])).font = font

    wb.save("poll_data.xlsx")


# color all votes.
def calculate_color(min_value, max_value, value):
    # 222,239,245 is the maximum color,
    # 35,158,208 is the minimum color.
    ratio = 0
    if min_value == max_value:
        ratio = 1
    else:
        ratio = (value - min_value) / (max_value - min_value)
    r = int((222 - 35) * (1 - ratio)) + 35
    g = int((239 - 158) * (1 - ratio)) + 158
    b = int((245 - 208) * (1 - ratio)) + 208
    return f'{r:02x}{g:02x}{b:02x}'


def convert_to_string(value):
    """ Convert value to string, handling different types carefully """
    if isinstance(value, (int, float)):
        # Convert numeric values to strings without scientific notation
        return format(value, 'f').rstrip('0').rstrip('.')
    return str(value)


# transfer from Excel to google sheet.
def hex_to_rgb(hex_color):
    """Convert hex color (e.g., 'FF0000') to RGB tuple (red, green, blue)."""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))


def transfer_sheet(from_name, to_name):
    # The name of the Excel file.
    excel_file = from_name
    # The id of the Google sheet.
    sheet_id = to_name

    # Credentials stuff.
    scope = ['https://spreadsheets.google.com/feeds',
             'https://www.googleapis.com/auth/drive',
             'https://www.googleapis.com/auth/spreadsheets']
    credentials = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)
    client = gspread.authorize(credentials)

    # Copying the Excel.
    df = pd.read_excel(excel_file)
    df.replace([np.inf, -np.inf], '', inplace=True)
    df.fillna('', inplace=True)
    df = df.applymap(convert_to_string)

    # Updating the Google sheet.
    sheet = client.open_by_key(sheet_id)
    worksheet = sheet.get_worksheet(0)
    data = df.values.tolist()
    worksheet.update('A2', data)

    # Build the service for the Google Sheets API
    service = build('sheets', 'v4', credentials=credentials)

    # Read the Excel file with formatting
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    requests = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
        for col_idx, cell in enumerate(row):
            if cell.fill.start_color.rgb is not None:
                rgb_color = 0
                if isinstance(cell.fill.start_color.rgb, openpyxl.styles.colors.RGB):
                    rgb_color = (0, 0, 0)
                else:
                    rgb_color = hex_to_rgb(cell.fill.start_color.rgb[2:])  # skip the first two characters 'FF'
                if rgb_color == (0, 0, 0):
                    rgb_color = (255, 255, 255)
                cell_format = {
                    'userEnteredFormat': {
                        'backgroundColor': {
                            'red': rgb_color[0] / 255.0,
                            'green': rgb_color[1] / 255.0,
                            'blue': rgb_color[2] / 255.0
                        }
                    }
                }

                requests.append({
                    'updateCells': {
                        'rows': {
                            'values': [{
                                'userEnteredFormat': cell_format['userEnteredFormat']
                            }]
                        },
                        'fields': 'userEnteredFormat.backgroundColor',
                        'start': {
                            'sheetId': worksheet.id,
                            'rowIndex': row_idx,
                            'columnIndex': col_idx
                        }
                    }
                })

    # Send batchUpdate request
    body = {
        'requests': requests
    }
    response = service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body=body).execute()

    print(f"Updated {len(response.get('replies', []))} cells with formatting.")
